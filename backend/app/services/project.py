"""项目管理服务模块。"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.config import settings
from app.models.project import (
    Project,
    ProjectCostRecord,
    ProjectDocumentFile,
    ProjectIssue,
    ProjectIssueImage,
    ProjectMilestone,
    ProjectPlan,
    ProjectRequirement,
)

UPLOAD_ROOT = Path(settings.UPLOAD_DIR)
ISSUE_IMAGE_DIR = UPLOAD_ROOT / "issues"
DOCUMENTS_ROOT = Path(settings.DOCUMENTS_DIR)
DOCUMENTS_FILES_ROOT = DOCUMENTS_ROOT / "files"

PLAN_SHEET_NAME = "项目开发计划"
PLAN_TEMPLATE_HEADERS = [
    "阶段",
    "一级任务",
    "二级任务（如有）",
    "依赖项\n（关联依赖项管理Sheet序列号）",
    "当前状态",
    "工期",
    "当前进度",
    "计划开始\n时间",
    "计划结束\n时间",
    "实际开始时间",
    "实际结束时间",
    "备注",
]
PLAN_STATUS_IMPORT_MAP = {
    "未开始": "待开始",
}
PLAN_STATUS_EXPORT_MAP = {
    "待开始": "未开始",
}


def list_projects(
    db: Session,
    status_filter: str | None = None,
    priority_filter: str | None = None,
    project_manager_filter: str | None = None,
    keyword: str | None = None,
) -> list[Project]:
    """获取项目列表并支持筛选。"""
    query = db.query(Project)

    if status_filter:
        query = query.filter(Project.status == status_filter)
    if priority_filter:
        query = query.filter(Project.priority == priority_filter)
    if project_manager_filter:
        query = query.filter(Project.project_manager == project_manager_filter)
    if keyword:
        fuzzy_keyword = f"%{keyword.strip()}%"
        query = query.filter(
            or_(
                Project.name.ilike(fuzzy_keyword),
                Project.description.ilike(fuzzy_keyword),
                Project.client_name.ilike(fuzzy_keyword),
            )
        )

    return query.order_by(Project.created_at.desc(), Project.id.desc()).all()


def get_project(db: Session, project_id: int) -> Project:
    """根据ID获取项目。"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"项目ID {project_id} 不存在",
        )
    return project


def create_project(
    db: Session,
    name: str,
    status: str = "规划中",
    priority: str = "中",
    description: str | None = None,
    announcement_markdown: str | None = None,
    docs_repo_url: str | None = None,
    docs_repo_branch: str | None = None,
    docs_repo_subpath: str | None = None,
    project_manager: str | None = None,
    client_name: str | None = None,
    git_url: str | None = None,
    start_date: date | None = None,
    planned_end_date: date | None = None,
    actual_end_date: date | None = None,
) -> Project:
    """创建项目。"""
    project = Project(
        name=name.strip(),
        status=status,
        priority=priority,
        description=description,
        announcement_markdown=announcement_markdown,
        docs_repo_url=docs_repo_url,
        docs_repo_branch=docs_repo_branch,
        docs_repo_subpath=docs_repo_subpath,
        project_manager=project_manager,
        client_name=client_name,
        git_url=git_url,
        start_date=start_date,
        planned_end_date=planned_end_date,
        actual_end_date=actual_end_date,
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


def update_project(db: Session, project_id: int, update_data: dict) -> Project:
    """更新项目。"""
    project = get_project(db, project_id)

    if "name" in update_data:
        project.name = update_data["name"].strip()
    if "status" in update_data:
        project.status = update_data["status"]
    if "priority" in update_data:
        project.priority = update_data["priority"]
    if "description" in update_data:
        project.description = update_data["description"]
    if "announcement_markdown" in update_data:
        project.announcement_markdown = update_data["announcement_markdown"]
    if "docs_repo_url" in update_data:
        project.docs_repo_url = update_data["docs_repo_url"]
    if "docs_repo_branch" in update_data:
        project.docs_repo_branch = update_data["docs_repo_branch"]
    if "docs_repo_subpath" in update_data:
        project.docs_repo_subpath = update_data["docs_repo_subpath"]
    if "project_manager" in update_data:
        project.project_manager = update_data["project_manager"]
    if "client_name" in update_data:
        project.client_name = update_data["client_name"]
    if "git_url" in update_data:
        project.git_url = update_data["git_url"]
    if "start_date" in update_data:
        project.start_date = update_data["start_date"]
    if "planned_end_date" in update_data:
        project.planned_end_date = update_data["planned_end_date"]
    if "actual_end_date" in update_data:
        project.actual_end_date = update_data["actual_end_date"]

    db.commit()
    db.refresh(project)
    return project


def delete_project(db: Session, project_id: int) -> None:
    """删除项目。"""
    project = get_project(db, project_id)
    for issue in project.issues:
        for image in issue.images:
            _delete_issue_image_file(image.file_path)
    for document in project.documents:
        _delete_project_document_file(document.file_path)
    db.delete(project)
    db.commit()


def list_milestones(db: Session, project_id: int) -> list[ProjectMilestone]:
    """获取项目里程碑列表。"""
    get_project(db, project_id)
    return (
        db.query(ProjectMilestone)
        .filter(ProjectMilestone.project_id == project_id)
        .order_by(ProjectMilestone.planned_date.asc().nullslast(), ProjectMilestone.created_at.asc())
        .all()
    )


def create_milestone(
    db: Session,
    project_id: int,
    name: str,
    description: str | None = None,
    planned_date: date | None = None,
    actual_date: date | None = None,
    progress_pct: int = 0,
    status: str = "待开始",
) -> ProjectMilestone:
    """创建里程碑。"""
    get_project(db, project_id)
    milestone = ProjectMilestone(
        project_id=project_id,
        name=name.strip(),
        description=description,
        planned_date=planned_date,
        actual_date=actual_date,
        progress_pct=progress_pct,
        status=status,
    )
    db.add(milestone)
    db.commit()
    db.refresh(milestone)
    return milestone


def update_milestone(db: Session, milestone_id: int, update_data: dict) -> ProjectMilestone:
    """更新里程碑。"""
    milestone = db.query(ProjectMilestone).filter(ProjectMilestone.id == milestone_id).first()
    if milestone is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"里程碑ID {milestone_id} 不存在",
        )

    if "name" in update_data:
        milestone.name = update_data["name"].strip()
    if "description" in update_data:
        milestone.description = update_data["description"]
    if "planned_date" in update_data:
        milestone.planned_date = update_data["planned_date"]
    if "actual_date" in update_data:
        milestone.actual_date = update_data["actual_date"]
    if "progress_pct" in update_data:
        milestone.progress_pct = update_data["progress_pct"]
    if "status" in update_data:
        milestone.status = update_data["status"]

    db.commit()
    db.refresh(milestone)
    return milestone


def delete_milestone(db: Session, milestone_id: int) -> None:
    """删除里程碑。"""
    milestone = db.query(ProjectMilestone).filter(ProjectMilestone.id == milestone_id).first()
    if milestone is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"里程碑ID {milestone_id} 不存在",
        )
    db.delete(milestone)
    db.commit()


def list_plans(db: Session, project_id: int) -> list[ProjectPlan]:
    """获取项目计划列表。"""
    get_project(db, project_id)
    return (
        db.query(ProjectPlan)
        .filter(ProjectPlan.project_id == project_id)
        .order_by(ProjectPlan.planned_start.asc().nullslast(), ProjectPlan.created_at.asc())
        .all()
    )


def get_plan(db: Session, plan_id: int) -> ProjectPlan:
    """根据ID获取项目计划。"""
    plan = db.query(ProjectPlan).filter(ProjectPlan.id == plan_id).first()
    if plan is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"计划ID {plan_id} 不存在",
        )
    return plan


def create_plan(
    db: Session,
    project_id: int,
    phase_name: str,
    primary_task: str | None = None,
    secondary_task: str | None = None,
    dependency: str | None = None,
    duration: str | None = None,
    progress_pct: int = 0,
    description: str | None = None,
    planned_start: date | None = None,
    planned_end: date | None = None,
    actual_start: date | None = None,
    actual_end: date | None = None,
    status: str = "待开始",
    assignee: str | None = None,
) -> ProjectPlan:
    """创建项目计划。"""
    get_project(db, project_id)
    plan = ProjectPlan(
        project_id=project_id,
        phase_name=phase_name.strip(),
        primary_task=_clean_optional_text(primary_task),
        secondary_task=_clean_optional_text(secondary_task),
        dependency=_clean_optional_text(dependency),
        duration=_clean_optional_text(duration),
        progress_pct=_coerce_progress_pct(progress_pct),
        description=description,
        planned_start=planned_start,
        planned_end=planned_end,
        actual_start=actual_start,
        actual_end=actual_end,
        status=status,
        assignee=assignee,
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    return plan


def update_plan(db: Session, plan_id: int, update_data: dict) -> ProjectPlan:
    """更新项目计划。"""
    plan = db.query(ProjectPlan).filter(ProjectPlan.id == plan_id).first()
    if plan is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"计划ID {plan_id} 不存在",
        )

    if "phase_name" in update_data:
        plan.phase_name = update_data["phase_name"].strip()
    if "primary_task" in update_data:
        plan.primary_task = _clean_optional_text(update_data["primary_task"])
    if "secondary_task" in update_data:
        plan.secondary_task = _clean_optional_text(update_data["secondary_task"])
    if "dependency" in update_data:
        plan.dependency = _clean_optional_text(update_data["dependency"])
    if "duration" in update_data:
        plan.duration = _clean_optional_text(update_data["duration"])
    if "progress_pct" in update_data:
        plan.progress_pct = _coerce_progress_pct(update_data["progress_pct"])
    if "description" in update_data:
        plan.description = update_data["description"]
    if "planned_start" in update_data:
        plan.planned_start = update_data["planned_start"]
    if "planned_end" in update_data:
        plan.planned_end = update_data["planned_end"]
    if "actual_start" in update_data:
        plan.actual_start = update_data["actual_start"]
    if "actual_end" in update_data:
        plan.actual_end = update_data["actual_end"]
    if "status" in update_data:
        plan.status = update_data["status"]
    if "assignee" in update_data:
        plan.assignee = update_data["assignee"]

    db.commit()
    db.refresh(plan)
    return plan


def delete_plan(db: Session, plan_id: int) -> None:
    """删除计划。"""
    plan = db.query(ProjectPlan).filter(ProjectPlan.id == plan_id).first()
    if plan is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"计划ID {plan_id} 不存在",
        )
    db.delete(plan)
    db.commit()


def import_plans_from_excel(
    db: Session,
    project_id: int,
    file: UploadFile,
) -> dict:
    """从 Excel 项目开发计划 sheet 导入计划记录。"""
    get_project(db, project_id)
    file_bytes = file.file.read()
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="导入文件为空",
        )

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="无法读取 Excel 文件，请确认文件格式为 .xlsx",
        ) from exc

    worksheet = (
        workbook[PLAN_SHEET_NAME]
        if PLAN_SHEET_NAME in workbook.sheetnames
        else workbook.active
    )
    header_row = _find_plan_header_row(worksheet)
    column_map = _resolve_plan_column_map(worksheet, header_row)

    created_plans: list[ProjectPlan] = []
    errors: list[str] = []
    skipped_count = 0
    current_phase: str | None = None
    current_primary_task: str | None = None

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        row_data = _read_plan_excel_row(worksheet, row_index, column_map)
        if not _has_plan_business_value(row_data):
            skipped_count += 1
            continue

        raw_phase = row_data.get("phase_name")
        raw_primary_task = row_data.get("primary_task")
        raw_secondary_task = row_data.get("secondary_task")

        if raw_phase:
            current_phase = raw_phase
            if not raw_primary_task:
                current_primary_task = None
        if raw_primary_task:
            current_primary_task = raw_primary_task

        phase_name = raw_phase or current_phase or raw_primary_task or raw_secondary_task
        if not phase_name:
            skipped_count += 1
            errors.append(f"第 {row_index} 行缺少阶段或任务名称，已跳过")
            continue

        try:
            plan = ProjectPlan(
                project_id=project_id,
                phase_name=phase_name,
                primary_task=raw_primary_task or current_primary_task,
                secondary_task=raw_secondary_task,
                dependency=row_data.get("dependency"),
                duration=row_data.get("duration"),
                progress_pct=_parse_progress_value(row_data.get("progress_pct")),
                description=row_data.get("description"),
                planned_start=_parse_excel_date_value(row_data.get("planned_start")),
                planned_end=_parse_excel_date_value(row_data.get("planned_end")),
                actual_start=_parse_excel_date_value(row_data.get("actual_start")),
                actual_end=_parse_excel_date_value(row_data.get("actual_end")),
                status=_normalize_import_status(row_data.get("status")),
                assignee=None,
            )
        except ValueError as exc:
            skipped_count += 1
            errors.append(f"第 {row_index} 行解析失败：{exc}")
            continue

        db.add(plan)
        created_plans.append(plan)

    if created_plans:
        db.commit()
        for plan in created_plans:
            db.refresh(plan)

    return {
        "created_count": len(created_plans),
        "skipped_count": skipped_count,
        "errors": errors,
        "plans": created_plans,
    }


def export_plans_to_excel(db: Session, project_id: int) -> tuple[str, bytes]:
    """导出项目计划为项目开发计划模板格式。"""
    project = get_project(db, project_id)
    plans = list_plans(db, project_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = PLAN_SHEET_NAME

    _build_plan_template_header(worksheet, project.name)
    for row_offset, plan in enumerate(plans, start=5):
        _write_plan_export_row(worksheet, row_offset, plan)

    _style_plan_template(worksheet, max_row=max(worksheet.max_row, 5))
    output = BytesIO()
    workbook.save(output)
    filename = f"{project.name}_项目开发计划.xlsx"
    return filename, output.getvalue()


def _find_plan_header_row(worksheet: Worksheet) -> int:
    """定位项目开发计划表头行。"""
    for row_index in range(1, min(worksheet.max_row, 20) + 1):
        values = [
            _normalize_header_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, min(worksheet.max_column, 24) + 1)
        ]
        if "阶段" in values and "一级任务" in values:
            return row_index

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="未找到项目开发计划表头，请确认包含“阶段”和“一级任务”列",
    )


def _resolve_plan_column_map(worksheet: Worksheet, header_row: int) -> dict[str, int]:
    """解析项目计划 Excel 字段列号。"""
    normalized_headers = {
        _normalize_header_text(worksheet.cell(row=header_row, column=column_index).value): column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    required_headers = {
        "阶段": "phase_name",
        "一级任务": "primary_task",
        "二级任务（如有）": "secondary_task",
        "当前状态": "status",
        "计划开始时间": "planned_start",
        "计划结束时间": "planned_end",
    }
    optional_headers = {
        "依赖项（关联依赖项管理Sheet序列号）": "dependency",
        "工期": "duration",
        "当前进度": "progress_pct",
        "实际开始时间": "actual_start",
        "实际结束时间": "actual_end",
        "备注": "description",
    }

    column_map: dict[str, int] = {}
    missing_headers: list[str] = []
    for header, field_name in required_headers.items():
        column_index = normalized_headers.get(header)
        if column_index is None:
            missing_headers.append(header)
        else:
            column_map[field_name] = column_index

    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"导入文件缺少必要列：{', '.join(missing_headers)}",
        )

    for header, field_name in optional_headers.items():
        column_index = normalized_headers.get(header)
        if column_index is not None:
            column_map[field_name] = column_index

    return column_map


def _read_plan_excel_row(
    worksheet: Worksheet,
    row_index: int,
    column_map: dict[str, int],
) -> dict[str, object]:
    """读取 Excel 中的一行项目计划数据。"""
    row_data: dict[str, object] = {}
    for field_name, column_index in column_map.items():
        row_data[field_name] = worksheet.cell(row=row_index, column=column_index).value

    return {
        "phase_name": _clean_optional_text(row_data.get("phase_name")),
        "primary_task": _clean_optional_text(row_data.get("primary_task")),
        "secondary_task": _clean_optional_text(row_data.get("secondary_task")),
        "dependency": _clean_optional_text(row_data.get("dependency")),
        "status": _clean_optional_text(row_data.get("status")),
        "duration": _clean_optional_text(row_data.get("duration")),
        "progress_pct": row_data.get("progress_pct"),
        "planned_start": row_data.get("planned_start"),
        "planned_end": row_data.get("planned_end"),
        "actual_start": row_data.get("actual_start"),
        "actual_end": row_data.get("actual_end"),
        "description": _clean_optional_text(row_data.get("description")),
    }


def _has_plan_business_value(row_data: dict[str, object]) -> bool:
    """判断 Excel 行是否包含计划业务数据。"""
    fields = (
        "phase_name",
        "primary_task",
        "secondary_task",
        "status",
        "duration",
        "progress_pct",
        "planned_start",
        "planned_end",
        "actual_start",
        "actual_end",
        "description",
    )
    return any(row_data.get(field_name) not in (None, "") for field_name in fields)


def _normalize_header_text(value: object) -> str:
    """规范化 Excel 表头文本。"""
    if value is None:
        return ""
    return str(value).replace("\n", "").replace(" ", "").strip()


def _normalize_import_status(value: object) -> str:
    """规范化导入计划状态。"""
    status_text = _clean_optional_text(value) or "待开始"
    return PLAN_STATUS_IMPORT_MAP.get(status_text, status_text)


def _export_plan_status(value: str) -> str:
    """转换导出计划状态。"""
    return PLAN_STATUS_EXPORT_MAP.get(value, value)


def _parse_excel_date_value(value: object) -> date | None:
    """解析 Excel 日期值。"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception as exc:
            raise ValueError(f"日期序列值 {value} 无法解析") from exc
    if isinstance(value, str):
        normalized = value.strip().replace("/", "-").replace(".", "-")
        if not normalized:
            return None
        try:
            return date.fromisoformat(normalized)
        except ValueError as exc:
            raise ValueError(f"日期 {value} 格式无法解析") from exc

    raise ValueError(f"日期值 {value} 类型不支持")


def _parse_progress_value(value: object) -> int:
    """解析 Excel 当前进度。"""
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        normalized = value.strip().replace("%", "")
        if not normalized:
            return 0
        number = float(normalized)
        return _coerce_progress_pct(number)
    if isinstance(value, (int, float)):
        if 0 <= float(value) <= 1:
            return _coerce_progress_pct(float(value) * 100)
        return _coerce_progress_pct(value)

    raise ValueError(f"进度值 {value} 类型不支持")


def _coerce_progress_pct(value: object) -> int:
    """将进度值限制在 0 到 100。"""
    if value is None or value == "":
        return 0
    progress = int(round(float(value)))
    return max(0, min(progress, 100))


def _clean_optional_text(value: object) -> str | None:
    """清理可选文本字段。"""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _build_plan_template_header(worksheet: Worksheet, project_name: str) -> None:
    """生成项目开发计划模板表头。"""
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=36)
    worksheet.cell(row=1, column=1, value=f"{project_name}开发计划")

    for column_index, header in enumerate(PLAN_TEMPLATE_HEADERS, start=1):
        worksheet.cell(row=2, column=column_index, value=header)

    current_year = date.today().year
    worksheet.merge_cells(start_row=2, start_column=13, end_row=2, end_column=36)
    worksheet.cell(row=2, column=13, value=f"{current_year}年")

    month_labels = ["6月", "7月", "8月", "9月", "10月", "11月"]
    for month_offset, month_label in enumerate(month_labels):
        start_column = 13 + month_offset * 4
        worksheet.merge_cells(
            start_row=3,
            start_column=start_column,
            end_row=3,
            end_column=start_column + 3,
        )
        worksheet.cell(row=3, column=start_column, value=month_label)
        for week_offset in range(4):
            worksheet.cell(row=4, column=start_column + week_offset, value=f"{week_offset + 1}w")


def _write_plan_export_row(worksheet: Worksheet, row_index: int, plan: ProjectPlan) -> None:
    """写入一行计划导出数据。"""
    values = [
        plan.phase_name,
        plan.primary_task,
        plan.secondary_task,
        plan.dependency,
        _export_plan_status(plan.status),
        plan.duration,
        (plan.progress_pct or 0) / 100,
        plan.planned_start,
        plan.planned_end,
        plan.actual_start,
        plan.actual_end,
        plan.description,
    ]
    for column_index, value in enumerate(values, start=1):
        worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.cell(row=row_index, column=7).number_format = "0%"
    for column_index in (8, 9, 10, 11):
        worksheet.cell(row=row_index, column=column_index).number_format = "yyyy-mm-dd"


def _style_plan_template(worksheet: Worksheet, max_row: int) -> None:
    """设置项目开发计划导出样式。"""
    worksheet.freeze_panes = "A5"
    widths = [14, 22, 38, 18, 12, 10, 10, 14, 14, 14, 14, 24]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    for column_index in range(13, 37):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 6

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    timeline_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_side = Side(style="thin", color="A6A6A6")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    worksheet.cell(row=1, column=1).font = Font(name="等线", size=18, bold=True)
    worksheet.cell(row=1, column=1).fill = title_fill
    worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 30

    for row_index in range(2, max_row + 1):
        for column_index in range(1, 37):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row_index == 2 and column_index <= 12:
                cell.fill = header_fill
                cell.font = Font(name="等线", size=9, bold=True)
            elif row_index in (2, 3, 4) and column_index >= 13:
                cell.fill = timeline_fill
                cell.font = Font(name="等线", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def list_requirements(db: Session, project_id: int) -> list[ProjectRequirement]:
    """获取项目需求列表。"""
    get_project(db, project_id)
    return (
        db.query(ProjectRequirement)
        .filter(ProjectRequirement.project_id == project_id)
        .order_by(ProjectRequirement.created_at.asc())
        .all()
    )


def create_requirement(
    db: Session,
    project_id: int,
    title: str,
    req_id: str | None = None,
    description: str | None = None,
    priority: str = "中",
    status: str = "待评审",
    owner: str | None = None,
) -> ProjectRequirement:
    """创建需求。"""
    get_project(db, project_id)
    requirement = ProjectRequirement(
        project_id=project_id,
        req_id=req_id,
        title=title.strip(),
        description=description,
        priority=priority,
        status=status,
        owner=owner,
    )
    db.add(requirement)
    db.commit()
    db.refresh(requirement)
    return requirement


def update_requirement(
    db: Session,
    requirement_id: int,
    update_data: dict,
) -> ProjectRequirement:
    """更新需求。"""
    requirement = (
        db.query(ProjectRequirement).filter(ProjectRequirement.id == requirement_id).first()
    )
    if requirement is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"需求ID {requirement_id} 不存在",
        )

    if "title" in update_data:
        requirement.title = update_data["title"].strip()
    if "req_id" in update_data:
        requirement.req_id = update_data["req_id"]
    if "description" in update_data:
        requirement.description = update_data["description"]
    if "priority" in update_data:
        requirement.priority = update_data["priority"]
    if "status" in update_data:
        requirement.status = update_data["status"]
    if "owner" in update_data:
        requirement.owner = update_data["owner"]

    db.commit()
    db.refresh(requirement)
    return requirement


def delete_requirement(db: Session, requirement_id: int) -> None:
    """删除需求。"""
    requirement = (
        db.query(ProjectRequirement).filter(ProjectRequirement.id == requirement_id).first()
    )
    if requirement is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"需求ID {requirement_id} 不存在",
        )
    db.delete(requirement)
    db.commit()


def list_issues(db: Session, project_id: int) -> list[ProjectIssue]:
    """获取项目问题列表。"""
    get_project(db, project_id)
    return (
        db.query(ProjectIssue)
        .filter(ProjectIssue.project_id == project_id)
        .order_by(ProjectIssue.created_at.desc())
        .all()
    )


def create_issue(
    db: Session,
    project_id: int,
    title: str,
    description: str | None = None,
    severity: str = "一般",
    status: str = "待处理",
    assignee: str | None = None,
    resolution: str | None = None,
) -> ProjectIssue:
    """创建问题。"""
    get_project(db, project_id)
    issue = ProjectIssue(
        project_id=project_id,
        title=title.strip(),
        description=description,
        severity=severity,
        status=status,
        assignee=assignee,
        resolution=resolution,
    )
    db.add(issue)
    db.commit()
    db.refresh(issue)
    return issue


def update_issue(db: Session, issue_id: int, update_data: dict) -> ProjectIssue:
    """更新问题。"""
    issue = db.query(ProjectIssue).filter(ProjectIssue.id == issue_id).first()
    if issue is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"问题ID {issue_id} 不存在",
        )

    if "title" in update_data:
        issue.title = update_data["title"].strip()
    if "description" in update_data:
        issue.description = update_data["description"]
    if "severity" in update_data:
        issue.severity = update_data["severity"]
    if "status" in update_data:
        issue.status = update_data["status"]
    if "assignee" in update_data:
        issue.assignee = update_data["assignee"]
    if "resolution" in update_data:
        issue.resolution = update_data["resolution"]

    db.commit()
    db.refresh(issue)
    return issue


def delete_issue(db: Session, issue_id: int) -> None:
    """删除问题。"""
    issue = db.query(ProjectIssue).filter(ProjectIssue.id == issue_id).first()
    if issue is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"问题ID {issue_id} 不存在",
        )
    for image in issue.images:
        _delete_issue_image_file(image.file_path)
    db.delete(issue)
    db.commit()


def upload_issue_images(
    db: Session,
    issue_id: int,
    files: list[UploadFile],
) -> list[ProjectIssueImage]:
    """为问题上传多张图片。"""
    issue = db.query(ProjectIssue).filter(ProjectIssue.id == issue_id).first()
    if issue is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"问题ID {issue_id} 不存在",
        )

    if not files:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="至少上传一张图片",
        )

    ISSUE_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    uploaded_images: list[ProjectIssueImage] = []

    for file in files:
        if not (file.content_type or "").startswith("image/"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"文件“{file.filename}”不是图片",
            )

        file_bytes = file.file.read()
        if not file_bytes:
            continue

        suffix = Path(file.filename or "").suffix or ".png"
        stored_file_name = f"{uuid4().hex}{suffix}"
        relative_path = Path("issues") / str(issue_id) / stored_file_name
        absolute_dir = UPLOAD_ROOT / "issues" / str(issue_id)
        absolute_dir.mkdir(parents=True, exist_ok=True)
        absolute_path = absolute_dir / stored_file_name
        absolute_path.write_bytes(file_bytes)

        image = ProjectIssueImage(
            issue_id=issue_id,
            file_name=stored_file_name,
            original_name=file.filename or stored_file_name,
            file_path=str(relative_path).replace("\\", "/"),
            content_type=file.content_type,
        )
        db.add(image)
        uploaded_images.append(image)

    db.commit()
    for image in uploaded_images:
        db.refresh(image)
    return uploaded_images


def delete_issue_image(db: Session, image_id: int) -> None:
    """删除问题图片。"""
    image = db.query(ProjectIssueImage).filter(ProjectIssueImage.id == image_id).first()
    if image is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"问题图片ID {image_id} 不存在",
        )

    _delete_issue_image_file(image.file_path)
    db.delete(image)
    db.commit()


def build_issue_image_url(file_path: str) -> str:
    """构造问题图片访问地址。"""
    normalized_path = file_path.replace("\\", "/").lstrip("/")
    return f"/api/uploads/{normalized_path}"


def _delete_issue_image_file(file_path: str) -> None:
    """删除本地问题图片文件。"""
    absolute_path = UPLOAD_ROOT / Path(file_path)
    if absolute_path.exists():
        absolute_path.unlink()


def list_cost_records(db: Session, project_id: int) -> list[ProjectCostRecord]:
    """获取项目成本记录列表。"""
    get_project(db, project_id)
    return (
        db.query(ProjectCostRecord)
        .filter(ProjectCostRecord.project_id == project_id)
        .order_by(
            ProjectCostRecord.occurred_on.desc().nullslast(),
            ProjectCostRecord.created_at.desc(),
        )
        .all()
    )


def create_cost_record(
    db: Session,
    project_id: int,
    title: str,
    record_type: str,
    amount: Decimal,
    person: str | None = None,
    occurred_on: date | None = None,
    category: str | None = None,
    description: str | None = None,
) -> ProjectCostRecord:
    """创建项目成本记录。"""
    get_project(db, project_id)
    record = ProjectCostRecord(
        project_id=project_id,
        title=title.strip(),
        record_type=record_type,
        amount=amount,
        person=person,
        occurred_on=occurred_on,
        category=category,
        description=description,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def update_cost_record(db: Session, cost_id: int, update_data: dict) -> ProjectCostRecord:
    """更新项目成本记录。"""
    record = db.query(ProjectCostRecord).filter(ProjectCostRecord.id == cost_id).first()
    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"成本记录ID {cost_id} 不存在",
        )

    if "title" in update_data:
        record.title = update_data["title"].strip()
    if "record_type" in update_data:
        record.record_type = update_data["record_type"]
    if "amount" in update_data:
        record.amount = update_data["amount"]
    if "person" in update_data:
        record.person = update_data["person"]
    if "occurred_on" in update_data:
        record.occurred_on = update_data["occurred_on"]
    if "category" in update_data:
        record.category = update_data["category"]
    if "description" in update_data:
        record.description = update_data["description"]

    db.commit()
    db.refresh(record)
    return record


def delete_cost_record(db: Session, cost_id: int) -> None:
    """删除项目成本记录。"""
    record = db.query(ProjectCostRecord).filter(ProjectCostRecord.id == cost_id).first()
    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"成本记录ID {cost_id} 不存在",
        )
    db.delete(record)
    db.commit()


def summarize_cost_records(records: list[ProjectCostRecord]) -> dict[str, float]:
    """汇总项目收支。"""
    total_expense = sum(float(record.amount) for record in records if record.record_type == "expense")
    total_income = sum(float(record.amount) for record in records if record.record_type == "income")
    return {
        "total_expense": round(total_expense, 2),
        "total_income": round(total_income, 2),
        "balance": round(total_income - total_expense, 2),
    }


def get_project_documents_state(db: Session, project_id: int) -> dict:
    """获取项目文档状态。"""
    project = get_project(db, project_id)
    records = (
        db.query(ProjectDocumentFile)
        .filter(ProjectDocumentFile.project_id == project_id)
        .order_by(ProjectDocumentFile.created_at.desc(), ProjectDocumentFile.id.desc())
        .all()
    )
    return {
        "project_id": project.id,
        "files": [
            {
                "id": record.id,
                "project_id": record.project_id,
                "name": record.file_name,
                "original_name": record.original_name,
                "directory": record.directory,
                "relative_path": record.file_path,
                "file_url": build_project_document_url(record.file_path),
                "content_type": record.content_type,
                "size": record.file_size,
                "modified_at": record.updated_at or record.created_at,
            }
            for record in records
        ],
    }


def upload_project_documents(
    db: Session,
    project_id: int,
    files: list[UploadFile],
    directory: str | None = None,
) -> list[ProjectDocumentFile]:
    """上传项目文档。"""
    get_project(db, project_id)
    if not files:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="至少上传一个文件",
        )

    normalized_directory = _normalize_document_directory(directory)
    target_dir = DOCUMENTS_FILES_ROOT / f"project-{project_id}"
    if normalized_directory:
      target_dir = target_dir / normalized_directory
    target_dir.mkdir(parents=True, exist_ok=True)
    uploaded: list[ProjectDocumentFile] = []

    for file in files:
        file_bytes = file.file.read()
        if not file_bytes:
            continue

        suffix = Path(file.filename or "").suffix
        stored_file_name = f"{uuid4().hex}{suffix}"
        relative_path = Path("files") / f"project-{project_id}"
        if normalized_directory:
            relative_path = relative_path / normalized_directory
        relative_path = relative_path / stored_file_name
        absolute_path = target_dir / stored_file_name
        absolute_path.write_bytes(file_bytes)

        document = ProjectDocumentFile(
            project_id=project_id,
            file_name=stored_file_name,
            original_name=file.filename or stored_file_name,
            directory=normalized_directory,
            file_path=str(relative_path).replace("\\", "/"),
            content_type=file.content_type,
            file_size=len(file_bytes),
        )
        db.add(document)
        uploaded.append(document)

    db.commit()
    for document in uploaded:
        db.refresh(document)
    return uploaded


def delete_project_document(db: Session, document_id: int) -> None:
    """删除项目文档。"""
    document = db.query(ProjectDocumentFile).filter(ProjectDocumentFile.id == document_id).first()
    if document is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"项目文档ID {document_id} 不存在",
        )

    _delete_project_document_file(document.file_path)
    db.delete(document)
    db.commit()


def build_project_document_url(relative_path: str) -> str:
    """构造项目文档访问地址。"""
    normalized_path = relative_path.replace("\\", "/").lstrip("/")
    return f"/api/documents-storage/{normalized_path}"


def _delete_project_document_file(file_path: str) -> None:
    """删除本地项目文档文件。"""
    absolute_path = DOCUMENTS_ROOT / Path(file_path)
    if absolute_path.exists():
        absolute_path.unlink()


def _normalize_document_directory(directory: str | None) -> str | None:
    """规范化文档目录并防止越界。"""
    if directory is None:
        return None

    normalized = directory.strip().replace("\\", "/").strip("/")
    if not normalized:
        return None

    candidate = Path(normalized)
    if candidate.is_absolute() or ".." in candidate.parts:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="目标目录不合法",
        )
    return normalized
